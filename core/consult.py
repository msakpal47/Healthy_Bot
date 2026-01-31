import os
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List
import csv
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors

STORAGE_DIR = Path(__file__).resolve().parents[1] / "storage"
DATA_DIR = Path(__file__).resolve().parents[1] / "Data"
CSV_PATH = DATA_DIR / "100_unique_diseases.csv"

def ensure_storage() -> None:
    STORAGE_DIR.mkdir(parents=True, exist_ok=True)

def _read_json(path: Path) -> List[Dict[str, Any]]:
    if path.exists():
        import json
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def _write_json(path: Path, data: List[Dict[str, Any]]) -> None:
    import json
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def save_patient(patient: Dict[str, Any]) -> None:
    ensure_storage()
    p = STORAGE_DIR / "patients.json"
    data = _read_json(p)
    data.append(patient)
    _write_json(p, data)

def save_consultation(record: Dict[str, Any]) -> Path:
    ensure_storage()
    p = STORAGE_DIR / "consultations.json"
    data = _read_json(p)
    data.append(record)
    _write_json(p, data)
    return p

def _load_diseases() -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    if CSV_PATH.exists():
        with open(CSV_PATH.as_posix(), newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for r in reader:
                if not r.get("Disease"):
                    continue
                rows.append(r)
    try:
        import openpyxl  # type: ignore
    except Exception:
        return rows
    for xl in DATA_DIR.glob("*.xlsx"):
        try:
            wb = openpyxl.load_workbook(xl, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = {h: i for i, h in enumerate(headers)}
            for r in ws.iter_rows(min_row=2):
                d = {h: str(r[idx[h]].value or "").strip() if h in idx else "" for h in headers}
                if d.get("Disease"):
                    rows.append(d)
        except Exception:
            continue
    return rows

def _match_disease(name: str, rows: List[Dict[str, str]]) -> Dict[str, str] | None:
    n = (name or "").strip().lower()
    if not n:
        return None
    # exact match only
    for r in rows:
        if r.get("Disease","").strip().lower() == n:
            return r
    return None

def _extract_field(text: str, key_words: List[str]) -> str:
    for kw in key_words:
        # simple pattern: line starting with keyword or contains keyword followed by colon
        m = re.search(rf"(?im)^{kw}\s*[:\-]\s*(.+)$", text)
        if m:
            return m.group(1).strip()
        m = re.search(rf"(?i){kw}\s*[:\-]\s*(.+?)(?:\n|$)", text)
        if m:
            return m.group(1).strip()
    return ""

def doctor_response(patient: Dict[str, Any], qa) -> Dict[str, Any]:
    # Prefer CSV dataset mapping if available
    diseases = _load_diseases()
    row = _match_disease(patient.get("disease",""), diseases)
    if row:
        adult = row.get("Adult Dose","") or row.get("Adult dose","")
        child = row.get("Child Dose","") or row.get("Child dose","")
        dose = adult if int(patient.get("age", 0)) >= 18 else (child or adult)
        return {
            "date": datetime.now().isoformat(),
            "patient_name": patient.get("name",""),
            "disease": row.get("Disease",""),
            "medicine": row.get("Medicine",""),
            "dose": dose,
            "tests": row.get("Tests",""),
            "warning": row.get("Warnings",""),
            "home_remedy": row.get("Home Care",""),
        }
    # Fallback to retrieval from PDFs (guard if db unavailable)
    query = f"{patient.get('disease','')} {patient.get('symptoms','')}".strip()
    docs = []
    try:
        if getattr(qa, "db", None):
            docs = qa.db.similarity_search(query or "general health", k=3)
    except Exception:
        docs = []
    merged = "\n\n".join(d.page_content for d in docs)
    medicine = _extract_field(merged, ["medicine", "medicines", "drug"])
    tests = _extract_field(merged, ["tests", "investigation", "diagnostics"])
    warning = _extract_field(merged, ["warning", "caution", "contraindication"])
    home = _extract_field(merged, ["home remedy", "home treatment", "self care"])
    dose_adult = _extract_field(merged, ["adult dose", "dose (adult)", "dosage (adult)"])
    dose_child = _extract_field(merged, ["child dose", "dose (child)", "dosage (child)"])
    dose = dose_adult if int(patient.get("age", 0)) >= 18 else dose_child or dose_adult
    # Heuristic defaults for common 'fever' presentation
    if "fever" in (patient.get("symptoms","").lower()):
        if not medicine or ("supportive" in medicine.lower()):
            medicine = "Paracetamol, ORS for hydration"
    if not tests and "fever" in (patient.get("symptoms","").lower()):
        tests = "CBC, Platelet count, NS1 antigen"
    if not home:
        home = "Drink plenty of fluids; Paracetamol only (NO NSAIDs)"
    if not warning:
        warning = "Bleeding; Severe abdominal pain; Persistent vomiting"
    derived_disease = patient.get("disease","") or ("Dengue Fever" if "fever" in (patient.get("symptoms","").lower()) else "")
    return {
        "date": datetime.now().isoformat(),
        "patient_name": patient.get("name",""),
        "disease": derived_disease,
        "medicine": medicine,
        "dose": dose,
        "tests": tests,
        "warning": warning,
        "home_remedy": home,
    }

def save_consultation_pdf(patient: Dict[str, Any], reply: Dict[str, Any]) -> Path:
    ensure_storage()
    out = STORAGE_DIR / f"consult_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(out.as_posix(), pagesize=A4,
                            leftMargin=36, rightMargin=36, topMargin=48, bottomMargin=36)
    content = []
    title = Paragraph("Patient Health Report", styles["Title"])
    content.append(title)
    content.append(Spacer(1, 8))
    info_rows = [
        ["Date", reply.get("date","")],
        ["Name", patient.get("name","")],
        ["Age", str(patient.get("age",""))],
        ["Gender", patient.get("gender","")],
        ["Severity", patient.get("severity","")],
        ["Duration", patient.get("duration","")],
        ["Symptoms", patient.get("symptoms","")],
    ]
    info_table = Table(info_rows, colWidths=[90, 400])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.Color(0.92,0.95,1)),
        ("TEXTCOLOR", (0,0), (-1,0), colors.Color(0.12,0.28,0.56)),
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.Color(0.98,0.98,0.98)]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.Color(0.85,0.88,0.92)),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    content.append(info_table)
    content.append(Spacer(1, 12))
    diag_rows = [
        ["Disease", reply.get("disease","")],
        ["Medicines", reply.get("medicine","")],
        ["Dose", reply.get("dose","")],
        ["Recommended Tests", reply.get("tests","")],
        ["Home Care", reply.get("home_remedy","")],
    ]
    diag_table = Table(diag_rows, colWidths=[140, 350])
    diag_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.Color(0.93,1,0.97)),
        ("TEXTCOLOR", (0,0), (-1,0), colors.Color(0.03,0.4,0.28)),
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.Color(0.98,0.98,0.98)]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.Color(0.85,0.88,0.92)),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    content.append(diag_table)
    content.append(Spacer(1, 10))
    warn_text = reply.get("warning","")
    if warn_text:
        warn_table = Table([[f"WARNING: {warn_text}"]], colWidths=[490])
        warn_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), colors.Color(1,0.93,0.85)),
            ("TEXTCOLOR", (0,0), (-1,-1), colors.Color(0.55,0.18,0.07)),
            ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 10),
            ("BOX", (0,0), (-1,-1), 0.8, colors.Color(0.98,0.75,0.45)),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
            ("RIGHTPADDING", (0,0), (-1,-1), 8),
            ("TOPPADDING", (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ]))
        content.append(warn_table)
    def deco(canvas, _doc):
        canvas.saveState()
        canvas.setFillColor(colors.Color(0.23,0.51,0.96))
        canvas.rect(0, _doc.pagesize[1]-34, _doc.pagesize[0], 4, fill=1, stroke=0)
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.Color(0.85,0.88,0.92))
        canvas.drawString(_doc.leftMargin, 20, "Smart Health Assistant • Confidential")
        canvas.setFillColor(colors.Color(0.85,0.88,0.92))
        canvas.drawRightString(_doc.pagesize[0]-_doc.rightMargin, 20, f"Page {_doc.page}")
        canvas.restoreState()
    doc.build(content, onFirstPage=deco, onLaterPages=deco)
    return out
